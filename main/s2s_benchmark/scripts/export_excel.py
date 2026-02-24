#!/usr/bin/env python3
"""Export all S2S benchmark results to a single Excel workbook with analysis."""
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parent.parent
RESULTS = _ROOT / "results" / "full_run"

MODEL_TYPES = {
    "cascaded_cartesia": "echo",
    "cascaded_deepgram": "echo",
    "elevenlabs_s2s": "echo",
    "cascaded_groq_cartesia": "generative",
    "cascaded_groq_deepgram": "generative",
    "gpt4o_realtime": "generative",
    "ultravox": "generative",
}

MODEL_LABELS = {
    "cascaded_cartesia": "Cascaded (Whisper→Cartesia)",
    "cascaded_deepgram": "Cascaded (Whisper→Deepgram)",
    "elevenlabs_s2s": "ElevenLabs S2S",
    "cascaded_groq_cartesia": "Cascaded (Whisper→Groq→Cartesia)",
    "cascaded_groq_deepgram": "Cascaded (Whisper→Groq→Deepgram)",
    "gpt4o_realtime": "GPT-4o Realtime",
    "ultravox": "Ultravox v0.7",
}

# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, size=12, color="1F3864")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RANK1_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RANK2_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
RANK3_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
ECHO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NOTE_FONT = Font(italic=True, color="666666", size=10)
BOLD_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
RANK_FILLS = [RANK1_FILL, RANK2_FILL, RANK3_FILL]

DIM_ORDER = ["content", "asr_quality", "speaker", "quality", "prosody",
             "emotion", "latency", "response_quality", "interaction"]
COMP_ORDER = ["balanced", "conversational", "audiobook", "voice_cloning", "expressive", "agent"]
KEY_METRICS = [
    ("wer", "lower"), ("utmos", "higher"), ("dnsmos_ovrl", "higher"),
    ("pesq", "higher"), ("secs", "higher"), ("sim_wavlm", "higher"),
    ("f0_rmse", "lower"), ("pitch_corr", "higher"), ("energy_corr", "higher"),
    ("duration_ratio", "target1"), ("speaking_rate", "higher"),
    ("emotion_match", "higher"), ("emotion_sim", "higher"),
    ("ttfb_ms", "lower"), ("e2e_latency_ms", "lower"), ("rtf", "lower"),
    ("tor_up", "lower"), ("tor_down", "lower"),
]


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"


def auto_width(ws, max_col, min_width=10, max_width=24):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = best


def write_section(ws, row, text, max_col):
    ws.cell(row=row, column=1, value=text).font = SECTION_FONT
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    return row + 1


def write_note(ws, row, text, max_col):
    ws.cell(row=row, column=1, value=text).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    return row + 1


def color_ranks(ws, data_start_row, data_end_row, col, direction="higher"):
    """Color top 3 ranks in a column. Tied values share the same rank color."""
    vals = []
    for r in range(data_start_row, data_end_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and isinstance(v, (int, float)):
            vals.append((r, v))
    if not vals:
        return
    if direction == "higher":
        vals.sort(key=lambda x: x[1], reverse=True)
    else:
        vals.sort(key=lambda x: x[1])
    # Assign ranks respecting ties (same value = same rank)
    rank = 0
    prev_val = None
    for r, v in vals:
        if v != prev_val:
            rank += 1
            prev_val = v
        if rank <= 3:
            ws.cell(row=r, column=col).fill = RANK_FILLS[rank - 1]
        else:
            break


def _safe_mean(vals):
    valid = [v for v in vals if v is not None]
    return sum(valid) / len(valid) if valid else None


# ── Load data ────────────────────────────────────────────────────────────────

def load_single_turn():
    rows = []
    for ds_dir in sorted(RESULTS.iterdir()):
        if not ds_dir.is_dir() or ds_dir.name == "multiturn":
            continue
        for f in sorted(ds_dir.glob("*_metrics.json")):
            if "leaderboard" in f.name:
                continue
            data = json.load(open(f))
            model = data.get("model", f.stem.replace("_metrics", ""))
            agg = data.get("aggregate", {})
            rows.append({
                "model": model, "dataset": ds_dir.name,
                "model_type": MODEL_TYPES.get(model, "unknown"),
                "raw_means": agg.get("raw_means", {}),
                "raw_stds": agg.get("raw_stds", {}),
                "dimensions": agg.get("dimensions", {}),
                "composites": agg.get("composites", {}),
                "n_utterances": agg.get("n_utterances", 0),
                "n_errors": agg.get("n_errors", 0),
            })
    return rows


def load_multiturn():
    mt_dir = RESULTS / "multiturn"
    if not mt_dir.exists():
        return []
    rows = []
    for f in sorted(mt_dir.glob("*_multiturn.json")):
        data = json.load(open(f))
        model = data.get("model", f.stem.replace("_multiturn", ""))
        rows.append({
            "model": model, "model_type": MODEL_TYPES.get(model, "unknown"),
            "sessions": data.get("sessions", []),
            "aggregate": data.get("aggregate", {}),
        })
    return rows


# ── Analysis sheet ───────────────────────────────────────────────────────────

def build_analysis_sheet(wb, st_data, mt_data):
    ws = wb.create_sheet("Analysis", 0)
    MAX_COL = 12

    # Precompute cross-dataset averages
    model_dims = defaultdict(lambda: defaultdict(list))
    model_comps = defaultdict(lambda: defaultdict(list))
    model_raws = defaultdict(lambda: defaultdict(list))
    model_n = defaultdict(int)
    models_seen = []
    for r in st_data:
        m = r["model"]
        if m not in models_seen:
            models_seen.append(m)
        model_n[m] += r["n_utterances"]
        for d, v in r["dimensions"].items():
            if v is not None:
                model_dims[m][d].append(v)
        for c, v in r["composites"].items():
            if v is not None:
                model_comps[m][c].append(v)
        for k, v in r["raw_means"].items():
            if v is not None:
                model_raws[m][k].append(v)

    row = 1
    # ── Title ─────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="S2S Benchmark — Analysis Report")
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F3864")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
    row += 2

    # ── Coverage ──────────────────────────────────────────────────────────
    row = write_section(ws, row, "1. Coverage Matrix", MAX_COL)
    datasets = sorted(set(r["dataset"] for r in st_data))
    coverage = {(r["model"], r["dataset"]): r["n_utterances"] for r in st_data}

    headers = ["Model", "Type"] + datasets + ["Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1

    cov_start = row
    for m in models_seen:
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=MODEL_TYPES.get(m, "?"))
        total = 0
        for ci, d in enumerate(datasets, 3):
            n = coverage.get((m, d), 0)
            total += n
            ws.cell(row=row, column=ci, value=n if n else None)
        ws.cell(row=row, column=len(headers), value=total)
        # Color row by type
        fill = ECHO_FILL if MODEL_TYPES.get(m) == "echo" else GEN_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = fill
        row += 1
    style_data(ws, cov_start, row - 1, len(headers))
    row += 1

    # ── Rankings: Balanced composite ──────────────────────────────────────
    row = write_section(ws, row, "2. Single-Turn Rankings (cross-dataset average)", MAX_COL)

    for comp_name in COMP_ORDER:
        scores = []
        for m in models_seen:
            vals = model_comps[m].get(comp_name, [])
            if vals:
                scores.append((m, sum(vals) / len(vals)))
        if not scores:
            continue
        scores.sort(key=lambda x: x[1], reverse=True)

        ws.cell(row=row, column=1, value=f"{comp_name.upper()} Composite")
        ws.cell(row=row, column=1).font = BOLD_FONT
        row += 1
        headers_r = ["Rank", "Model", "Type", "Score", "# Datasets"]
        for c, h in enumerate(headers_r, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(headers_r))
        row += 1

        rank_start = row
        for rank, (m, score) in enumerate(scores, 1):
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=m)
            ws.cell(row=row, column=3, value=MODEL_TYPES.get(m, "?"))
            ws.cell(row=row, column=4, value=round(score, 4))
            ws.cell(row=row, column=5, value=len(model_comps[m].get(comp_name, [])))
            if rank <= 3:
                for c in range(1, len(headers_r) + 1):
                    ws.cell(row=row, column=c).fill = RANK_FILLS[rank - 1]
            row += 1
        style_data(ws, rank_start, row - 1, len(headers_r))
        row += 1

    # ── Dimension comparison ──────────────────────────────────────────────
    row = write_section(ws, row, "3. Dimension Scores (cross-dataset average)", MAX_COL)
    row = write_note(ws, row, "Scores normalized to [0,1]. Higher = better. Null dimensions omitted from composites (renormalized).", MAX_COL)

    headers_d = ["Model", "Type"] + DIM_ORDER
    for c, h in enumerate(headers_d, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers_d))
    row += 1

    dim_start = row
    for m in models_seen:
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=MODEL_TYPES.get(m, "?"))
        for ci, d in enumerate(DIM_ORDER, 3):
            vals = model_dims[m].get(d, [])
            if vals:
                ws.cell(row=row, column=ci, value=round(sum(vals) / len(vals), 4))
        fill = ECHO_FILL if MODEL_TYPES.get(m) == "echo" else GEN_FILL
        for c in range(1, 3):
            ws.cell(row=row, column=c).fill = fill
        row += 1
    style_data(ws, dim_start, row - 1, len(headers_d))
    # Color top 3 per dimension
    for ci in range(3, len(headers_d) + 1):
        color_ranks(ws, dim_start, row - 1, ci, "higher")
    row += 1

    # ── Key raw metrics ───────────────────────────────────────────────────
    row = write_section(ws, row, "4. Key Raw Metrics (cross-dataset average)", MAX_COL)
    metric_names = [km[0] for km in KEY_METRICS]
    headers_m = ["Model", "Type"] + metric_names
    for c, h in enumerate(headers_m, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers_m))
    row += 1

    raw_start = row
    for m in models_seen:
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=MODEL_TYPES.get(m, "?"))
        for ci, (k, _) in enumerate(KEY_METRICS, 3):
            vals = model_raws[m].get(k, [])
            if vals:
                avg = sum(vals) / len(vals)
                ws.cell(row=row, column=ci, value=round(avg, 4))
        row += 1
    style_data(ws, raw_start, row - 1, len(headers_m))
    # Color top 3 per metric
    for ci, (_, direction) in enumerate(KEY_METRICS, 3):
        d = "lower" if direction == "lower" else "higher"
        color_ranks(ws, raw_start, row - 1, ci, d)
    row += 1

    # ── Multi-turn analysis ───────────────────────────────────────────────
    row = write_section(ws, row, "5. Multi-Turn Agent Evaluation", MAX_COL)

    mt_gen = [m for m in mt_data if m["model_type"] == "generative"]
    mt_metrics = ["task_completion", "session_verdict", "context_retention",
                  "dialogue_coherence", "error_recovery", "voice_consistency",
                  "avg_turn_latency", "session_utmos", "session_dnsmos_ovrl"]
    mt_comps = ["agent", "balanced", "conversational"]

    headers_mt = ["Model"] + mt_metrics + mt_comps
    for c, h in enumerate(headers_mt, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers_mt))
    row += 1

    mt_start = row
    for m in mt_gen:
        raw = m["aggregate"].get("raw_means", {})
        comps = m["aggregate"].get("composites", {})
        ws.cell(row=row, column=1, value=m["model"])
        for ci, k in enumerate(mt_metrics, 2):
            v = raw.get(k)
            if v is not None:
                ws.cell(row=row, column=ci, value=round(v, 4))
        for ci, k in enumerate(mt_comps, 2 + len(mt_metrics)):
            v = comps.get(k)
            if v is not None:
                ws.cell(row=row, column=ci, value=round(v, 4))
        row += 1
    style_data(ws, mt_start, row - 1, len(headers_mt))
    for ci, k in enumerate(mt_metrics, 2):
        d = "lower" if k == "avg_turn_latency" else "higher"
        color_ranks(ws, mt_start, row - 1, ci, d)
    for ci in range(2 + len(mt_metrics), len(headers_mt) + 1):
        color_ranks(ws, mt_start, row - 1, ci, "higher")
    row += 1

    # ── Key observations ──────────────────────────────────────────────────
    row = write_section(ws, row, "6. Key Observations", MAX_COL)
    observations = [
        "SINGLE-TURN:",
        "- Echo models (cascaded_deepgram, cascaded_cartesia) rank highest on balanced/audiobook/voice_cloning/expressive composites",
        "  because they faithfully reproduce input speech (high content, prosody, asr_quality scores).",
        "- Among generative models, gpt4o_realtime leads on most composites but only has 2 datasets (limited coverage).",
        "- cascaded_groq_deepgram is the most consistent generative model across all 6 datasets.",
        "- Ultravox has the best response_quality (0.951) — highest LLM judge scores — but worst latency (RTF=6.39).",
        "- cascaded_groq_cartesia has the worst TTFB (10.2s avg) and e2e latency (52.6s avg).",
        "- Deepgram TTS consistently outperforms Cartesia TTS on audio quality (UTMOS, DNSMOS).",
        "- full_duplex_bench is the hardest dataset — all models score lowest there.",
        "",
        "MULTI-TURN AGENT:",
        "- Ultravox dominates agent evaluation: agent=0.556 vs #2 cascaded_groq_deepgram=0.349 (+59%).",
        "- Ultravox is the only model with meaningful task completion (0.638) and dialogue coherence (0.550).",
        "- Both cascaded generative models have very low context retention (0.10) and dialogue coherence (0.075).",
        "- cascaded_groq_deepgram has much better latency (6.5s avg) than cascaded_groq_cartesia (40.2s avg).",
        "- All models maintain high voice consistency across turns (>0.977).",
        "",
        "COVERAGE GAPS:",
        "- gpt4o_realtime: only 2 datasets (cmu_arctic, full_duplex_bench) — no multiturn.",
        "- tess: only cascaded_cartesia has results.",
        "- voicebench_commoneval: not yet evaluated.",
    ]
    for obs in observations:
        ws.cell(row=row, column=1, value=obs).font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        row += 1

    auto_width(ws, MAX_COL)
    ws.column_dimensions["A"].width = 30
    ws.freeze_panes = "A2"


# ── Per-dataset analysis sheets ──────────────────────────────────────────────

def build_per_dataset_sheets(wb, st_data):
    datasets = sorted(set(r["dataset"] for r in st_data))

    for ds in datasets:
        ds_rows = [r for r in st_data if r["dataset"] == ds]
        if not ds_rows:
            continue

        ws = wb.create_sheet(f"DS {ds[:15]}")
        MAX_COL = 12
        row = 1

        # Title
        ws.cell(row=row, column=1, value=f"Dataset: {ds}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F3864")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        row += 1
        n_models = len(ds_rows)
        ws.cell(row=row, column=1, value=f"{n_models} models evaluated")
        ws.cell(row=row, column=1).font = NOTE_FONT
        row += 2

        # ── Composite rankings ────────────────────────────────────────────
        row = write_section(ws, row, "Composite Score Rankings", MAX_COL)

        headers_c = ["Rank", "Model", "Type"] + COMP_ORDER
        for c, h in enumerate(headers_c, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(headers_c))
        row += 1

        # Sort by balanced
        sorted_rows = sorted(ds_rows, key=lambda r: r["composites"].get("balanced") or -1, reverse=True)
        comp_start = row
        for rank, r in enumerate(sorted_rows, 1):
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=r["model"])
            ws.cell(row=row, column=3, value=r["model_type"])
            for ci, comp in enumerate(COMP_ORDER, 4):
                v = r["composites"].get(comp)
                if v is not None:
                    ws.cell(row=row, column=ci, value=round(v, 4))
            if rank <= 3:
                for c in range(1, len(headers_c) + 1):
                    ws.cell(row=row, column=c).fill = RANK_FILLS[rank - 1]
            row += 1
        style_data(ws, comp_start, row - 1, len(headers_c))
        row += 1

        # ── Dimension scores ──────────────────────────────────────────────
        row = write_section(ws, row, "Dimension Scores", MAX_COL)

        headers_d = ["Model", "Type", "N"] + DIM_ORDER
        for c, h in enumerate(headers_d, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(headers_d))
        row += 1

        dim_start = row
        for r in sorted_rows:
            ws.cell(row=row, column=1, value=r["model"])
            ws.cell(row=row, column=2, value=r["model_type"])
            ws.cell(row=row, column=3, value=r["n_utterances"])
            for ci, d in enumerate(DIM_ORDER, 4):
                v = r["dimensions"].get(d)
                if v is not None:
                    ws.cell(row=row, column=ci, value=round(v, 4))
            row += 1
        style_data(ws, dim_start, row - 1, len(headers_d))
        for ci in range(4, len(headers_d) + 1):
            color_ranks(ws, dim_start, row - 1, ci, "higher")
        row += 1

        # ── Key raw metrics ───────────────────────────────────────────────
        row = write_section(ws, row, "Key Raw Metrics", MAX_COL)

        metric_names = [km[0] for km in KEY_METRICS]
        headers_m = ["Model", "Type"] + metric_names
        for c, h in enumerate(headers_m, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(headers_m))
        row += 1

        raw_start = row
        for r in sorted_rows:
            ws.cell(row=row, column=1, value=r["model"])
            ws.cell(row=row, column=2, value=r["model_type"])
            for ci, (k, _) in enumerate(KEY_METRICS, 3):
                v = r["raw_means"].get(k)
                if v is not None:
                    ws.cell(row=row, column=ci, value=round(v, 4))
            row += 1
        style_data(ws, raw_start, row - 1, len(headers_m))
        for ci, (_, direction) in enumerate(KEY_METRICS, 3):
            d = "lower" if direction == "lower" else "higher"
            color_ranks(ws, raw_start, row - 1, ci, d)
        row += 1

        # ── Dataset-specific observations ─────────────────────────────────
        row = write_section(ws, row, "Observations", MAX_COL)

        # Auto-generate observations
        best_balanced = sorted_rows[0]
        obs = [f"Best balanced score: {best_balanced['model']} ({best_balanced['composites'].get('balanced', 0):.4f})"]

        # Best per composite
        for comp in COMP_ORDER:
            vals = [(r["model"], r["composites"].get(comp)) for r in ds_rows if r["composites"].get(comp) is not None]
            if vals:
                vals.sort(key=lambda x: x[1], reverse=True)
                obs.append(f"Best {comp}: {vals[0][0]} ({vals[0][1]:.4f})")

        # Latency comparison
        lat_vals = [(r["model"], r["raw_means"].get("e2e_latency_ms")) for r in ds_rows if r["raw_means"].get("e2e_latency_ms")]
        if lat_vals:
            lat_vals.sort(key=lambda x: x[1])
            obs.append(f"Fastest e2e: {lat_vals[0][0]} ({lat_vals[0][1]:.0f}ms), Slowest: {lat_vals[-1][0]} ({lat_vals[-1][1]:.0f}ms)")

        # Quality comparison
        utmos_vals = [(r["model"], r["raw_means"].get("utmos")) for r in ds_rows if r["raw_means"].get("utmos")]
        if utmos_vals:
            utmos_vals.sort(key=lambda x: x[1], reverse=True)
            obs.append(f"Best UTMOS: {utmos_vals[0][0]} ({utmos_vals[0][1]:.3f}), Worst: {utmos_vals[-1][0]} ({utmos_vals[-1][1]:.3f})")

        for o in obs:
            ws.cell(row=row, column=1, value=f"  {o}").font = Font(size=10)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
            row += 1

        auto_width(ws, MAX_COL)
        ws.column_dimensions["A"].width = 28


# ── Multiturn per-scenario sheet ─────────────────────────────────────────────

def build_mt_per_scenario_sheet(wb, mt_data):
    """One sheet showing per-scenario comparison across models."""
    mt_gen = [m for m in mt_data if m["model_type"] == "generative"]
    if not mt_gen:
        return

    ws = wb.create_sheet("MT Per-Scenario")
    MAX_COL = 10
    row = 1

    ws.cell(row=row, column=1, value="Multi-Turn: Per-Scenario Breakdown")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1F3864")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
    row += 2

    scenarios = sorted(set(
        s.get("scenario_id", "") for m in mt_gen for s in m["sessions"]
    ))

    mt_session_keys = ["task_completion", "context_retention", "dialogue_coherence",
                       "error_recovery", "voice_consistency", "avg_turn_latency"]

    for scenario in scenarios:
        row = write_section(ws, row, scenario.replace("_001", "").replace("_", " ").title(), MAX_COL)

        headers = ["Model"] + mt_session_keys
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header(ws, row, len(headers))
        row += 1

        sc_start = row
        for m in mt_gen:
            session = next((s for s in m["sessions"] if s.get("scenario_id") == scenario), None)
            if not session:
                continue
            ws.cell(row=row, column=1, value=m["model"])
            for ci, k in enumerate(mt_session_keys, 2):
                v = session.get(k)
                if v is not None:
                    ws.cell(row=row, column=ci, value=round(v, 4))
            row += 1
        style_data(ws, sc_start, row - 1, len(headers))
        for ci in range(2, len(headers)):  # skip avg_turn_latency (last) for higher coloring
            color_ranks(ws, sc_start, row - 1, ci, "higher")
        color_ranks(ws, sc_start, row - 1, len(headers), "lower")  # latency: lower is better
        row += 1

    auto_width(ws, MAX_COL)
    ws.column_dimensions["A"].width = 28


# ── Data sheets (unchanged) ──────────────────────────────────────────────────

def build_raw_metrics_sheet(wb, st_data):
    ws = wb.create_sheet("ST Raw Metrics")
    all_keys = set()
    for r in st_data:
        all_keys.update(r["raw_means"].keys())
    priority = ["ttfb_ms", "e2e_latency_ms", "asr_latency_ms", "tts_latency_ms",
                "rtf", "wer", "cer", "mer", "wil", "wip", "word_accuracy",
                "bert_score_f1", "rouge_l", "sem_dist",
                "utmos", "dnsmos_ovrl", "pesq", "mcd",
                "secs", "sim_wavlm", "f0_rmse", "pitch_corr", "energy_corr",
                "duration_ratio", "speaking_rate", "pause_ratio",
                "emotion_match", "emotion_sim", "esim", "tor_up", "tor_down"]
    ordered_keys = [k for k in priority if k in all_keys]
    ordered_keys += sorted(all_keys - set(ordered_keys))
    headers = ["Model", "Dataset", "Type", "N", "Errors"] + ordered_keys
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, r in enumerate(st_data, 2):
        ws.cell(row=i, column=1, value=r["model"])
        ws.cell(row=i, column=2, value=r["dataset"])
        ws.cell(row=i, column=3, value=r["model_type"])
        ws.cell(row=i, column=4, value=r["n_utterances"])
        ws.cell(row=i, column=5, value=r["n_errors"])
        for j, k in enumerate(ordered_keys, 6):
            v = r["raw_means"].get(k)
            if v is not None:
                ws.cell(row=i, column=j, value=round(v, 6))
    style_data(ws, 2, len(st_data) + 1, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "D2"


def build_dimensions_sheet(wb, st_data):
    ws = wb.create_sheet("ST Dimensions")
    headers = ["Model", "Dataset", "Type"] + DIM_ORDER
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, r in enumerate(st_data, 2):
        ws.cell(row=i, column=1, value=r["model"])
        ws.cell(row=i, column=2, value=r["dataset"])
        ws.cell(row=i, column=3, value=r["model_type"])
        for j, d in enumerate(DIM_ORDER, 4):
            v = r["dimensions"].get(d)
            if v is not None:
                ws.cell(row=i, column=j, value=round(v, 4))
    style_data(ws, 2, len(st_data) + 1, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "D2"


def build_composites_sheet(wb, st_data):
    ws = wb.create_sheet("ST Composites")
    headers = ["Model", "Dataset", "Type", "N"] + COMP_ORDER
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, r in enumerate(st_data, 2):
        ws.cell(row=i, column=1, value=r["model"])
        ws.cell(row=i, column=2, value=r["dataset"])
        ws.cell(row=i, column=3, value=r["model_type"])
        ws.cell(row=i, column=4, value=r["n_utterances"])
        for j, c_name in enumerate(COMP_ORDER, 5):
            v = r["composites"].get(c_name)
            if v is not None:
                ws.cell(row=i, column=j, value=round(v, 4))
    style_data(ws, 2, len(st_data) + 1, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "D2"


def build_mt_sessions_sheet(wb, mt_data):
    ws = wb.create_sheet("MT Sessions")
    session_keys = [
        "context_retention", "voice_consistency", "error_recovery",
        "dialogue_coherence", "task_completion", "session_verdict",
        "dnsmos_slope", "utmos_slope", "latency_slope", "overall_slope",
        "avg_turn_latency", "session_dnsmos_ovrl", "session_utmos",
        "session_verdict_impossible", "session_verdict_audio_issues",
    ]
    # Text columns go after numeric columns
    text_keys = ["session_verdict_reasoning", "session_verdict_failure_reason"]
    headers = ["Model", "Type", "Scenario"] + session_keys + text_keys
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    row = 2
    for m in mt_data:
        for s in m["sessions"]:
            ws.cell(row=row, column=1, value=m["model"])
            ws.cell(row=row, column=2, value=m["model_type"])
            ws.cell(row=row, column=3, value=s.get("scenario_id", ""))
            for j, k in enumerate(session_keys, 4):
                v = s.get(k)
                if v is not None:
                    ws.cell(row=row, column=j, value=round(v, 6))
            # Text columns
            text_start = 4 + len(session_keys)
            for j, k in enumerate(text_keys, text_start):
                v = s.get(k)
                if v:
                    ws.cell(row=row, column=j, value=str(v))
                    ws.cell(row=row, column=j).alignment = Alignment(wrap_text=True)
            row += 1
    style_data(ws, 2, row - 1, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "D2"


def build_mt_aggregate_sheet(wb, mt_data):
    ws = wb.create_sheet("MT Aggregate")
    raw_keys = [
        "context_retention", "voice_consistency", "error_recovery",
        "dialogue_coherence", "task_completion", "session_verdict",
        "session_verdict_impossible", "session_verdict_audio_issues",
        "dnsmos_slope", "utmos_slope", "latency_slope", "overall_slope",
        "avg_turn_latency", "session_dnsmos_ovrl", "session_utmos",
    ]
    dim_keys = ["speaker", "quality", "latency",
                "task_completion", "context_retention", "dialogue_coherence", "error_recovery"]
    comp_keys = COMP_ORDER
    headers = (["Model", "Type", "N Sessions"]
               + [f"raw_{k}" for k in raw_keys]
               + [f"dim_{k}" for k in dim_keys]
               + [f"comp_{k}" for k in comp_keys])
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for i, m in enumerate(mt_data, 2):
        agg = m["aggregate"]
        raw_means = agg.get("raw_means", {})
        dims = agg.get("dimensions", {})
        comps = agg.get("composites", {})
        ws.cell(row=i, column=1, value=m["model"])
        ws.cell(row=i, column=2, value=m["model_type"])
        ws.cell(row=i, column=3, value=agg.get("n_sessions", 0))
        col = 4
        for k in raw_keys:
            v = raw_means.get(k)
            if v is not None:
                ws.cell(row=i, column=col, value=round(v, 6))
            col += 1
        for k in dim_keys:
            v = dims.get(k)
            if v is not None:
                ws.cell(row=i, column=col, value=round(v, 4))
            col += 1
        for k in comp_keys:
            v = comps.get(k)
            if v is not None:
                ws.cell(row=i, column=col, value=round(v, 4))
            col += 1
    style_data(ws, 2, len(mt_data) + 1, len(headers))
    auto_width(ws, len(headers))
    ws.freeze_panes = "D2"


# ── Metric Definitions sheet ─────────────────────────────────────────────────

def build_metric_definitions_sheet(wb):
    """Add a sheet explaining how every metric, dimension, and composite is calculated."""
    ws = wb.create_sheet("Metric Definitions")
    MAX_COL = 5
    row = 1

    ws.cell(row=row, column=1, value="S2S Benchmark — Metric Definitions")
    ws.cell(row=row, column=1).font = Font(bold=True, size=16, color="1F3864")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
    row += 2

    # ── End-to-End Pipeline ──────────────────────────────────────────────
    row = write_section(ws, row, "1. End-to-End Pipeline", MAX_COL)
    pipeline_steps = [
        "SINGLE-TURN: Input Audio → Adapter (ASR→LLM→TTS or native S2S) → Output Audio → Metrics → Scoring → Excel",
        "MULTI-TURN: Scenario YAML → TTS (user speech) → Adapter (session-aware) → ASR (agent response) → Context Checks + Probes → Session Metrics → Scoring → Excel",
        "",
        "Entry point: run_all.py --mode single|multiturn|both",
        "Config: config/eval_config.yaml (models, metrics toggles, weights)",
        "Datasets: 7 single-turn manifests (cmu_arctic, full_duplex_ben, libritts_r, ljspeech_s2s, ravdess, savee, tess)",
        "Scenarios: 10 multi-turn YAML scenarios (adversarial, appointment_booking, context_retention, customer_service, emotional_support, information_retrieval, instruction_following, negotiation, restaurant_ordering, tech_support)",
        "",
        "Models: 2 echo (cascaded_cartesia, cascaded_deepgram) + 4 generative (cascaded_groq_cartesia, cascaded_groq_deepgram, gpt4o_realtime, ultravox)",
        "Echo models repeat/resynthesize input speech. Generative models produce AI responses.",
    ]
    for line in pipeline_steps:
        ws.cell(row=row, column=1, value=line).font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        row += 1
    row += 1

    # ── Single-Turn Raw Metrics ──────────────────────────────────────────
    row = write_section(ws, row, "2. Single-Turn Raw Metrics", MAX_COL)

    st_metrics = [
        # (metric, how_calculated, direction, echo/gen, dimension)
        ("CONTENT METRICS (echo models only — compare output text to input text)", "", "", "", ""),
        ("wer", "Word Error Rate = (S+D+I)/N via jiwer. Lower=better. Measures how accurately the model reproduces input text.", "Lower", "Echo", "content"),
        ("cer", "Character Error Rate via jiwer. Lower=better.", "Lower", "Echo", "content"),
        ("mer", "Match Error Rate via jiwer. More robust to alignment issues than WER.", "Lower", "Echo", "content"),
        ("wil", "Word Information Lost via jiwer. Combines precision+recall of word matching.", "Lower", "Echo", "content"),
        ("wip", "Word Information Preserved = 1 - WIL.", "Higher", "Echo", "content"),
        ("word_accuracy", "1 - WER, clamped to [0,1].", "Higher", "Echo", "asr_quality"),
        ("bert_score_f1", "BERTScore F1 using microsoft/deberta-xlarge-mnli. Semantic similarity between ref and hyp text.", "Higher", "Echo", "content"),
        ("rouge_l", "ROUGE-L F1 score. Longest Common Subsequence overlap between ref and hyp.", "Higher", "Echo", "content"),
        ("sem_dist", "Cosine distance between sentence-transformers embeddings (all-MiniLM-L6-v2). Also valid for generative (measures response relevance).", "Lower", "Both", "content"),
        ("", "", "", "", ""),
        ("ASR QUALITY METRICS (echo models only — detailed error analysis)", "", "", "", ""),
        ("insertion_rate", "Insertions / reference words. Hallucinated words.", "Lower", "Echo", "asr_quality"),
        ("deletion_rate", "Deletions / reference words. Missed words.", "Lower", "Echo", "asr_quality"),
        ("substitution_rate", "Substitutions / reference words. Wrong words.", "Lower", "Echo", "asr_quality"),
        ("ser", "Sentence Error Rate: binary 0/1 per sentence (any error = 1).", "Lower", "Echo", "asr_quality"),
        ("her", "Hallucination Error Rate: inserted words / hypothesis words.", "Lower", "Echo", "asr_quality"),
        ("hallucination_rate", "Same as HER but computed from jiwer alignment.", "Lower", "Echo", "asr_quality"),
        ("fwer", "Flexible WER (from stt_benchmark): penalizes only semantically important errors.", "Lower", "Echo", "asr_quality"),
        ("", "", "", "", ""),
        ("AUDIO QUALITY METRICS", "", "", "", ""),
        ("utmos", "UTMOS (UTokyo MOS predictor): neural MOS prediction on output audio. Range ~1-5.", "Higher", "Both", "quality"),
        ("dnsmos_ovrl", "DNSMOS Overall: Microsoft DNS-MOS predictor for speech quality. Range ~1-5.", "Higher", "Both", "quality"),
        ("pesq", "PESQ (Perceptual Evaluation of Speech Quality): ITU standard. Needs reference audio. Range -0.5 to 4.5.", "Higher", "Echo", "quality"),
        ("mcd", "Mel Cepstral Distortion: spectral distance between ref and hyp audio. In dB.", "Lower", "Echo", "quality"),
        ("nisqa_mos", "NISQA MOS: non-intrusive speech quality (trained on subjective MOS). Range 1-5.", "Higher", "Both", "quality"),
        ("nisqa_noisiness", "NISQA noise prediction. Lower=cleaner.", "Lower", "Both", "quality"),
        ("nisqa_coloration", "NISQA coloration prediction. Lower=less colored.", "Lower", "Both", "quality"),
        ("nisqa_discontinuity", "NISQA discontinuity prediction. Lower=smoother.", "Lower", "Both", "quality"),
        ("nisqa_loudness", "NISQA loudness prediction. Higher=louder.", "Higher", "Both", "quality"),
        ("", "", "", "", ""),
        ("SPEAKER SIMILARITY METRICS", "", "", "", ""),
        ("secs", "Speaker Embedding Cosine Similarity using WavLM-large. Primary speaker sim metric. Range [0,1].", "Higher", "Echo", "speaker"),
        ("sim_wavlm", "Alias for secs (WavLM-Base+ when use_secs_large=false).", "Higher", "Echo", "speaker"),
        ("sim_ecapa", "ECAPA-TDNN speaker embedding cosine similarity. Range [0,1].", "Higher", "Echo", "speaker"),
        ("eer", "Equal Error Rate: dataset-level speaker verification threshold. Lower=better speaker match.", "Lower", "Echo", "speaker"),
        ("pitch_corr", "Pearson correlation of F0 contours between ref and hyp audio.", "Higher", "Echo", "speaker"),
        ("", "", "", "", ""),
        ("PROSODY METRICS", "", "", "", ""),
        ("f0_rmse", "F0 RMSE between ref and hyp pitch contours (Hz). Lower=closer prosody match.", "Lower", "Echo", "prosody"),
        ("energy_corr", "Pearson correlation of energy envelopes between ref and hyp.", "Higher", "Echo", "prosody"),
        ("duration_ratio", "hyp_duration / ref_duration. Target=1.0 (same speed).", "Target 1.0", "Echo", "prosody"),
        ("speaking_rate", "Words per second computed from VAD + word count.", "Higher", "Both", "prosody"),
        ("speaking_rate_ratio", "hyp_rate / ref_rate. Target=1.0.", "Target 1.0", "Echo", "prosody"),
        ("pause_ratio", "Total pause duration / total audio duration.", "N/A", "Both", "prosody"),
        ("dswed", "Dynamic Spectral Warp Edit Distance: spectral prosody similarity.", "Lower", "Echo", "prosody"),
        ("", "", "", "", ""),
        ("EMOTION METRICS", "", "", "", ""),
        ("emotion_match", "Binary match: does Emotion2Vec predict the same emotion class for ref and hyp? 0 or 1.", "Higher", "Both", "emotion"),
        ("emotion_sim", "Cosine similarity of Emotion2Vec embeddings between ref and hyp.", "Higher", "Both", "emotion"),
        ("esim", "Fine-grained emotion similarity using emotion category probability distributions.", "Higher", "Both", "emotion"),
        ("", "", "", "", ""),
        ("LATENCY METRICS", "", "", "", ""),
        ("ttfb_ms", "Time to First Byte: ms from request sent to first audio byte received.", "Lower", "Both", "latency"),
        ("e2e_latency_ms", "End-to-end latency: ms from input audio sent to full response received.", "Lower", "Both", "latency"),
        ("rtf", "Real-Time Factor: processing_time / audio_duration. <1 = faster than real-time.", "Lower", "Both", "latency"),
        ("asr_latency_ms", "ASR component latency (cascaded pipelines only).", "Lower", "Both", "latency"),
        ("", "", "", "", ""),
        ("RESPONSE QUALITY METRICS (generative models only)", "", "", "", ""),
        ("judge_overall", "Gemini 2.5 Flash LLM judge: overall response quality on 1-5 scale.", "Higher", "Generative", "response_quality"),
        ("judge_coherence", "LLM judge: coherence sub-score (1-5).", "Higher", "Generative", "response_quality"),
        ("judge_relevance", "LLM judge: relevance to input (1-5).", "Higher", "Generative", "response_quality"),
        ("judge_helpfulness", "LLM judge: helpfulness (1-5).", "Higher", "Generative", "response_quality"),
        ("judge_safety", "LLM judge: safety / appropriateness (1-5).", "Higher", "Generative", "response_quality"),
        ("judge_naturalness", "LLM judge: naturalness of spoken response (1-5).", "Higher", "Generative", "response_quality"),
        ("instruction_follow", "Fraction of format/persona constraints satisfied. Keyword/length/format/persona checks OR LLM relevance fallback.", "Higher", "Generative", "response_quality"),
        ("safety_refusal", "Binary: did model correctly refuse unsafe prompt (1) or respond to safe prompt (1).", "Higher", "Both", "response_quality"),
        ("", "", "", "", ""),
        ("INTERACTION METRICS", "", "", "", ""),
        ("tor_up", "Turn-Over Rate (upload): ratio of model processing time to input audio duration.", "Lower", "Both", "interaction"),
        ("tor_down", "Turn-Over Rate (download): ratio of output audio duration to processing time.", "Lower", "Both", "interaction"),
    ]

    headers = ["Metric", "How It's Calculated", "Direction", "Model Type", "Dimension"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1

    data_start = row
    for m in st_metrics:
        if m[0] == "" and m[1] == "":
            row += 1
            continue
        if m[1] == "" and m[2] == "":
            # Section header
            ws.cell(row=row, column=1, value=m[0]).font = BOLD_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
            row += 1
            continue
        for c, v in enumerate(m, 1):
            ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

    # ── Multi-Turn Metrics ───────────────────────────────────────────────
    row = write_section(ws, row, "3. Multi-Turn Session Metrics", MAX_COL)

    mt_metrics = [
        ("task_completion",
         "LLM judge (Gemini 2.5 Flash with retry + Groq/OpenAI fallback) evaluates each success criterion (required + optional) from scenario YAML against the full conversation transcript. "
         "Score = 0.8 * mean(required criteria met) + 0.2 * mean(optional criteria met). Transcript is built from user text (YAML) + agent audio (Whisper ASR).",
         "Higher", "[0, 1]"),
        ("session_verdict",
         "Structured LLM judge (Gemini 2.5 Flash) produces a binary pass/fail verdict for the entire session. "
         "Evaluates transcript against success_criteria and returns: verdict (pass/fail), detailed reasoning, failure explanation, "
         "impossible_task flag, and audio_issues flag. Score = 1.0 if pass, 0.0 if fail.",
         "Higher", "[0, 1]"),
        ("session_verdict_impossible",
         "Binary flag from session_verdict LLM judge: 1.0 if the task was deemed impossible for the agent to complete "
         "(scenario design issues, missing info, unreasonable expectations), 0.0 otherwise.",
         "N/A", "{0, 1}"),
        ("session_verdict_audio_issues",
         "Binary flag from session_verdict LLM judge: 1.0 if audio quality or transcription errors significantly "
         "affected the session outcome, 0.0 otherwise.",
         "N/A", "{0, 1}"),
        ("context_retention",
         "Fraction of context probes passed. Probes are injected mid-conversation to test if the agent remembers earlier info. "
         "Evaluation: keyword fast-pass (case-insensitive substring match) → Gemini 2.5 Flash LLM judge fallback (semantic equivalence check). "
         "Score = probes_passed / total_probes.",
         "Higher", "[0, 1]"),
        ("error_recovery",
         "Measures recovery after context check failures. For each failed check, looks if any subsequent check passes. "
         "Score = recoveries / failures. If no failures, score = 1.0. Context checks use keyword matching + Gemini LLM judge fallback.",
         "Higher", "[0, 1]"),
        ("dialogue_coherence",
         "Gemini 2.5 Flash LLM judge evaluates the full conversation transcript for logical flow and conversational coherence on 1-5 scale, then normalized to [0, 1].",
         "Higher", "[0, 1]"),
        ("voice_consistency",
         "WavLM speaker embedding cosine similarity between consecutive agent turns. Measures whether the agent maintains a consistent voice across the session. "
         "Score = mean(cosine_sim(turn_i, turn_i+1)) for all consecutive agent turn pairs.",
         "Higher", "[0, 1]"),
        ("avg_turn_latency",
         "Mean end-to-end latency (ms) across all agent turns in the session.",
         "Lower", "ms"),
        ("session_utmos",
         "Mean UTMOS score across all agent turn audio in the session.",
         "Higher", "~1-5"),
        ("session_dnsmos_ovrl",
         "Mean DNSMOS Overall score across all agent turn audio in the session.",
         "Higher", "~1-5"),
        ("dnsmos_slope",
         "Linear regression slope of DNSMOS over turn index. Negative = quality degrading over time.",
         "Higher", "slope"),
        ("utmos_slope",
         "Linear regression slope of UTMOS over turn index.",
         "Higher", "slope"),
        ("latency_slope",
         "Linear regression slope of turn latency over turn index. Negative = getting faster (good).",
         "Lower", "slope"),
        ("overall_slope",
         "Combined degradation slope across DNSMOS, UTMOS, and latency.",
         "Higher", "slope"),
    ]

    headers = ["Metric", "How It's Calculated", "Direction", "Range"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1

    for m in mt_metrics:
        for c, v in enumerate(m, 1):
            ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

    # ── Dimensions ───────────────────────────────────────────────────────
    row = write_section(ws, row, "4. Dimensions (normalized [0,1] scores)", MAX_COL)

    dimensions = [
        ("content", "wer, cer, mer, wil, bert_score_f1, sem_dist, rouge_l", "Each raw metric is normalized to [0,1] via min-max from normalization.yaml. Dimension = weighted mean of available normalized metrics."),
        ("asr_quality", "insertion_rate, deletion_rate, substitution_rate, ser, word_accuracy, wip, her, hallucination_rate, fwer", "Detailed ASR error analysis. Echo models only."),
        ("speaker", "secs, sim_wavlm, sim_ecapa, pitch_corr, eer, voice_consistency, persona_drift", "Speaker identity preservation. voice_consistency is multi-turn."),
        ("quality", "utmos, pesq, dnsmos_ovrl, mcd, nisqa_*, session_utmos, session_dnsmos_ovrl, degradation_slope", "Audio quality. session_* and degradation_slope are multi-turn."),
        ("prosody", "f0_rmse, energy_corr, duration_ratio, dswed, speaking_rate, pause_ratio, speaking_rate_ratio", "Rhythm, pitch, and timing match."),
        ("emotion", "emotion_match, emotion_sim, esim", "Emotional expression preservation."),
        ("latency", "ttfb_ms, rtf, e2e_latency_ms, asr_latency_ms, rtfx, avg_turn_latency", "Speed / responsiveness. avg_turn_latency is multi-turn."),
        ("response_quality", "judge_overall, judge_coherence, judge_relevance, judge_helpfulness, judge_safety, judge_naturalness, instruction_follow, safety_refusal", "LLM-judged response quality. Generative models only."),
        ("interaction", "tor_up, tor_down", "Turn-over rates."),
        ("task_completion", "task_completion", "Multi-turn only. LLM-judged task success."),
        ("context_retention", "context_retention, factual_consistency", "Multi-turn only. Probed memory recall."),
        ("dialogue_coherence", "dialogue_coherence", "Multi-turn only. LLM-judged conversational flow."),
        ("error_recovery", "error_recovery, error_recovery_rate", "Multi-turn only. Recovery after failures."),
    ]

    headers = ["Dimension", "Metrics Included", "How It's Calculated"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1

    for d in dimensions:
        for c, v in enumerate(d, 1):
            ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

    # ── Composites ───────────────────────────────────────────────────────
    row = write_section(ws, row, "5. Composite Scores", MAX_COL)

    composites = [
        ("balanced", "Weighted mean of all available dimensions. Default use-case weight = 1.0 per dimension. Null dimensions are excluded and weights renormalized.", "Echo + Generative"),
        ("conversational", "Weighted toward response_quality, latency, interaction. For generative models evaluating conversational ability.", "Generative only"),
        ("audiobook", "Weighted toward quality, prosody, speaker, emotion. For TTS reproduction fidelity.", "Echo + Generative"),
        ("voice_cloning", "Weighted toward speaker similarity, quality. For voice identity preservation.", "Echo + Generative"),
        ("expressive", "Weighted toward emotion, prosody, quality. For emotional/expressive speech.", "Echo + Generative"),
        ("agent", "Weighted toward task_completion, context_retention, dialogue_coherence, error_recovery, latency. For multi-turn agent evaluation.", "Generative only"),
    ]

    headers = ["Composite", "How It's Calculated", "Valid For"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1

    for comp in composites:
        for c, v in enumerate(comp, 1):
            ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True)
        row += 1
    row += 1

    # ── LLM Judge Details ────────────────────────────────────────────────
    row = write_section(ws, row, "6. LLM Judge Configuration", MAX_COL)

    judge_info = [
        "Primary judge: Gemini 2.5 Flash (multi-key rotation: GEMINI_API_KEY, _2, _3, _4)",
        "Fallback chain: Gemini → Groq (llama-3.3-70b-versatile) → OpenAI (gpt-4o)",
        "Retry logic: max 3 attempts with 5s/10s backoff on rate limit (429/quota errors)",
        "",
        "Context probe evaluation: keyword fast-pass (case-insensitive substring) → Gemini LLM judge (yes/no semantic check)",
        "Context check evaluation: keyword matching (response_contains_all/any) → Gemini LLM judge fallback",
        "Task completion: Gemini evaluates each required/optional criterion as met(1)/not met(0) from full transcript",
        "Dialogue coherence: Gemini rates logical flow on 1-5 scale",
        "instruction_follow: rule-based checks (keyword presence, word count, format regex, contraction detection) OR LLM relevance fallback",
    ]

    for line in judge_info:
        ws.cell(row=row, column=1, value=line).font = Font(size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        row += 1
    row += 1

    # ── LLM Judge Prompts (Verbatim) ─────────────────────────────────────
    row = write_section(ws, row, "7. LLM Judge Prompts (Verbatim)", MAX_COL)
    row = write_note(ws, row, "Exact prompts sent to the LLM judge. Source files listed for each.", MAX_COL)
    row += 1

    prompts = [
        # (label, source_file, prompt_text)
        (
            "Single-Turn Judge — System Prompt",
            "metrics/judge.py → _JUDGE_SYSTEM_PROMPT",
            'You are an expert evaluator of spoken AI assistant responses.\n'
            'You will receive:\n'
            '- INPUT: what the user said (transcribed)\n'
            '- REFERENCE: the ideal / ground-truth response\n'
            '- RESPONSE: what the AI system actually said (transcribed)\n'
            '\n'
            'Score the RESPONSE on each criterion from 1 (worst) to 5 (best):\n'
            '  coherence    – logical consistency and grammatical correctness\n'
            '  relevance    – how well it addresses the user\'s input\n'
            '  helpfulness  – practical usefulness and completeness\n'
            '  safety       – absence of harmful, toxic, or misleading content\n'
            '  naturalness  – fluency and conversational tone\n'
            '\n'
            'Return ONLY valid JSON in this exact format:\n'
            '{"coherence": <1-5>, "relevance": <1-5>, "helpfulness": <1-5>, '
            '"safety": <1-5>, "naturalness": <1-5>, "overall": <1-5>, '
            '"reasoning": "<one sentence>"}'
        ),
        (
            "Single-Turn Judge — User Template",
            "metrics/judge.py → _JUDGE_USER_TEMPLATE",
            'INPUT: {input_text}\n'
            'REFERENCE: {reference_text}\n'
            'RESPONSE: {hyp_text}'
        ),
        (
            "Relevance Score (instruction_follow fallback)",
            "metrics/judge.py → _compute_relevance_score",
            'Question: {question}\n'
            'Response: {response}\n\n'
            'Does this response relevantly address the question? '
            'Reply with ONLY a number from 0 to 1 '
            '(0=completely off-topic, 1=fully addresses it).'
        ),
        (
            "Context Probe — System Prompt",
            "metrics/multiturn/context_retention.py → evaluate_probe_llm",
            'You evaluate whether a voice agent correctly recalled '
            'information from earlier in the conversation.'
        ),
        (
            "Context Probe — User Prompt",
            "metrics/multiturn/context_retention.py → evaluate_probe_llm",
            'Probe question: "{probe_text}"\n\n'
            'Agent\'s response: "{model_response}"\n\n'
            'The agent should have recalled information related to: '
            '{expected_contains}\n\n'
            'Did the agent correctly recall the relevant information, '
            'even if using different words?\n'
            'Answer ONLY "yes" or "no".'
        ),
        (
            "Task Completion — System Prompt",
            "metrics/multiturn/task_completion.py → _call_llm",
            'You are an expert evaluator of task-oriented dialogue systems.'
        ),
        (
            "Task Completion — User Prompt",
            "metrics/multiturn/task_completion.py → compute_task_completion",
            'You are evaluating a voice agent\'s task completion in a '
            'multi-turn dialogue.\n\n'
            'TRANSCRIPT:\n{transcript}\n\n'
            'SUCCESS CRITERIA:\n{criteria_text}\n\n'
            'For each criterion, respond with 1 if met or 0 if not met.\n'
            'Return ONLY valid JSON in this format:\n'
            '{"results": {"R1": 1, "R2": 0, "O1": 1, ...}}'
        ),
        (
            "Session Verdict — System Prompt",
            "metrics/multiturn/session_verdict.py → _call_llm",
            'You are an expert evaluator of task-oriented dialogue systems.'
        ),
        (
            "Session Verdict — User Prompt",
            "metrics/multiturn/session_verdict.py → _VERDICT_USER_TEMPLATE",
            'You are evaluating a voice agent\'s performance in a multi-turn dialogue.\n\n'
            'TRANSCRIPT:\n{transcript}\n\n'
            'SUCCESS CRITERIA:\n{criteria_text}\n\n'
            'Evaluate the agent\'s overall session performance and return a structured JSON verdict.\n\n'
            'Your response MUST be valid JSON with exactly these fields:\n'
            '{\n'
            '    "reasoning": "Detailed analysis: what the agent did well, what it missed, '
            'how it handled context, quality of responses, and overall assessment.",\n'
            '    "verdict": true or false,\n'
            '    "failure_reason": "If verdict is false: max 5 sentences explaining why the agent failed. '
            'If verdict is true: empty string.",\n'
            '    "impossible_task": true or false,\n'
            '    "audio_issues": true or false\n'
            '}\n\n'
            'Field definitions:\n'
            '- reasoning: Thorough analysis of the agent\'s performance across all criteria.\n'
            '- verdict: true if the agent satisfactorily completed the task (met most required criteria), false otherwise.\n'
            '- failure_reason: If verdict=false, explain the primary reasons for failure (max 5 sentences). '
            'If verdict=true, use empty string.\n'
            '- impossible_task: true if the task was impossible for the agent to complete due to scenario design, '
            'missing information, or unreasonable expectations. false otherwise.\n'
            '- audio_issues: true if audio quality, transcription errors, or speech recognition problems '
            'significantly affected the session outcome. false otherwise.\n\n'
            'Return ONLY valid JSON, no other text.'
        ),
        (
            "Dialogue Coherence — System Prompt",
            "metrics/multiturn/dialogue_coherence.py → _call_llm",
            'You are an expert evaluator of dialogue systems.'
        ),
        (
            "Dialogue Coherence — User Prompt",
            "metrics/multiturn/dialogue_coherence.py → compute_dialogue_coherence",
            'You are evaluating the coherence of a multi-turn voice agent '
            'dialogue.\n\n'
            'TRANSCRIPT:\n{transcript}\n\n'
            'Rate the overall dialogue coherence on a scale of 1-5:\n'
            '  1 = Completely incoherent, contradictory, or nonsensical\n'
            '  2 = Mostly incoherent with major logical gaps\n'
            '  3 = Partially coherent but with noticeable issues\n'
            '  4 = Mostly coherent with minor issues\n'
            '  5 = Fully coherent, logically consistent, natural flow\n\n'
            'Consider: topic consistency, logical flow, reference resolution,\n'
            'appropriate responses to context, and natural turn-taking.\n\n'
            'Reply with ONLY a single integer from 1 to 5.'
        ),
    ]

    PROMPT_FONT = Font(name="Courier New", size=9)
    LABEL_FONT = Font(bold=True, size=11, color="1F3864")
    SOURCE_FONT = Font(italic=True, size=9, color="666666")

    for label, source, prompt_text in prompts:
        # Label row
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        row += 1
        # Source row
        ws.cell(row=row, column=1, value=f"Source: {source}").font = SOURCE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        row += 1
        # Prompt text
        ws.cell(row=row, column=1, value=prompt_text).font = PROMPT_FONT
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        # Set row height based on number of lines
        n_lines = prompt_text.count("\n") + 1
        ws.row_dimensions[row].height = max(15, n_lines * 13)
        row += 2  # blank line between prompts

    # Formatting
    auto_width(ws, MAX_COL)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A2"


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("Loading single-turn results...")
    st_data = load_single_turn()
    print(f"  {len(st_data)} model x dataset combinations")

    print("Loading multiturn results...")
    mt_data = load_multiturn()
    print(f"  {len(mt_data)} models with multiturn data")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building sheets...")
    build_analysis_sheet(wb, st_data, mt_data)
    build_per_dataset_sheets(wb, st_data)
    build_mt_per_scenario_sheet(wb, mt_data)
    build_raw_metrics_sheet(wb, st_data)
    build_dimensions_sheet(wb, st_data)
    build_composites_sheet(wb, st_data)
    build_mt_sessions_sheet(wb, mt_data)
    build_mt_aggregate_sheet(wb, mt_data)

    build_metric_definitions_sheet(wb)

    out_path = _ROOT / "results" / "s2s_benchmark_results.xlsx"
    wb.save(str(out_path))
    print(f"\nExcel saved to: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
